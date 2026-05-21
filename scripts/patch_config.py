"""
patch_config.py
---------------
3PL Inbound Onboarding Automation — MuleSoft YAML patcher + Translation table generator.

Usage:
    python patch_config.py --input sample_onboarding_input.json
                           --yaml  dev.yaml
                           --out-yaml  dev_patched.yaml
                           --out-translations translations_new_rows.xlsx

What it does:
    1. Reads the onboarding JSON (filled-in by the requestor).
    2. Validates all mandatory fields.
    3. Appends the new country block to the YAML (preserving all existing content + comments).
    4. Generates the new Translation Table rows as Excel + CSV.
    5. Prints a human-readable summary of every change made.
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

# openpyxl for Excel output — install with: pip install openpyxl
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── All valid transaction codes ──────────────────────────────────────────────
VALID_TX_CODES = {
    "sal_008": "SHIPMENT",
    "sal_011": "SALES_RETURN_RECEIPT",
    "pur_019": "RECADV",
    "man_001": "FINISHED_GOOD",
    "inv_002": "RECLASSMENT",
    "inv_003": "INVENTORY",
    "inv_004": "STOCK_ADJUSTMENT",
    "wms_004": "TO_SHIP_CONFIRMATION",
    "wms_005": "TO_REC_CONFIRMATION",
}

# ── Mandatory fields (dot-path) ──────────────────────────────────────────────
REQUIRED_FIELDS = [
    "country_key",
    "nav.protocol",
    "nav.host",
    "nav.port",
    "nav.username",
    "nav.domain",
    "nav.company",
    "nav.service",
    "nav.soap_port",
    "nav.soap_path",
    "nav.routing_code",
    "nav.transaction_types",
    "translation.receiver_name",
    "translation.message_types",
]

# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

def _get(d: dict, dotpath: str):
    """Traverse nested dict using dot-separated path."""
    parts = dotpath.split(".")
    cur = d
    for p in parts:
        if not isinstance(cur, dict) or p not in cur:
            return None
        cur = cur[p]
    return cur


def validate(data: dict) -> list[str]:
    """Return list of error strings; empty list = valid."""
    errors = []
    for field in REQUIRED_FIELDS:
        val = _get(data, field)
        if val is None or val == "" or val == []:
            errors.append(f"Missing mandatory field: '{field}'")

    # country_key must be lowercase letters only
    ck = data.get("country_key", "")
    if ck and not re.fullmatch(r"[a-z]+", ck):
        errors.append(
            f"'country_key' must be lowercase letters only (no spaces/numbers). Got: '{ck}'"
        )

    # At least one transaction type must be enabled
    tx = _get(data, "nav.transaction_types") or {}
    enabled = [k for k, v in tx.items() if isinstance(v, dict) and v.get("enabled")]
    if not enabled:
        errors.append("No transaction types are enabled. Enable at least one in nav.transaction_types.")

    # receiver_name must not be empty
    rn = _get(data, "translation.receiver_name") or ""
    if rn.strip() == "" or "PLACEHOLDER" in rn.upper():
        errors.append("'translation.receiver_name' looks like a placeholder — set the real BTP ClientID.")

    return errors


# ────────────────────────────────────────────────────────────────────────────
# YAML block generator
# ────────────────────────────────────────────────────────────────────────────

def build_yaml_block(data: dict) -> str:
    """
    Generate the new country YAML block in the exact style used by the existing file:
      - Top-level key: no indentation
      - Properties: 3-space indent
      - transactiontype sub-entries: 5-space indent
    """
    nav = data["nav"]
    country_key = data["country_key"]
    comment = data.get("partner_comment", "").strip()

    lines = []
    if comment:
        lines.append(f"#{comment}")

    lines.append(f"{country_key}:")
    lines.append(f'   protocol: "{nav["protocol"]}"')
    lines.append(f'   host: "{nav["host"]}"')
    lines.append(f'   port: "{nav["port"]}"')
    lines.append(f'   username: "{nav["username"]}"')
    lines.append(f'   domain: "{nav["domain"]}"')
    lines.append(f'   company: "{nav["company"]}"')
    lines.append(f'   service: "{nav["service"]}"')
    lines.append(f'   soap.port: "{nav["soap_port"]}"')
    lines.append(f'   soap.path: "{nav["soap_path"]}"')
    lines.append(f'   routing.code: "{nav["routing_code"]}"')
    lines.append( '   transactiontype:')

    tx = nav.get("transaction_types", {})
    # Preserve canonical order from VALID_TX_CODES
    for code in VALID_TX_CODES:
        entry = tx.get(code, {})
        if isinstance(entry, dict) and entry.get("enabled"):
            type_name = entry.get("type", VALID_TX_CODES[code])
            lines.append(f'     {code}: "{type_name}"')

    return "\n".join(lines) + "\n"


# ────────────────────────────────────────────────────────────────────────────
# YAML patcher
# ────────────────────────────────────────────────────────────────────────────

def patch_yaml(yaml_path: str, country_key: str, new_block: str) -> str:
    """
    Read the YAML file, check the country doesn't already exist,
    append the new block at the end, return the updated content.
    """
    with open(yaml_path, "r", encoding="utf-8") as f:
        content = f.read()

    # Guard: don't add a duplicate
    # A country key appears as a line starting with exactly '{key}:'
    pattern = re.compile(rf"^{re.escape(country_key)}\s*:", re.MULTILINE)
    if pattern.search(content):
        raise ValueError(
            f"Country key '{country_key}' already exists in {yaml_path}. "
            "Remove it manually before re-running, or choose a different key."
        )

    separator = "\n" if content.endswith("\n") else "\n\n"
    return content + separator + new_block


# ────────────────────────────────────────────────────────────────────────────
# Translation row generator
# ────────────────────────────────────────────────────────────────────────────

def build_translation_rows(data: dict) -> list[dict]:
    """
    Build rows matching the SharePoint/BTP translation table schema:
    ReceiverName | MessageType | TranslationSchema | TranslationName | ValueFrom | ValueTo
    """
    rows = []
    t = data["translation"]
    receiver = t["receiver_name"]
    today = datetime.today().strftime("%-m/%-d/%Y %H:%M") if os.name != "nt" else datetime.today().strftime("%#m/%#d/%Y %H:%M")
    created_by = data.get("created_by", "Automation Script")

    for msg_type in t.get("message_types", []):
        # sourceDestinationCombination rows
        for combo in t.get("source_destination_combinations", []):
            rows.append({
                "ReceiverName":      receiver,
                "MessageType":       msg_type,
                "TranslationSchema": "INTERNAL_TO_EXTERNAL",
                "TranslationName":   "sourceDestinationCombination",
                "ValueFrom":         combo["value_from"],
                "ValueTo":           combo["value_to"],
                "Created By":        created_by,
                "Modified By":       created_by,
                "Modified":          today,
                "Item Type":         "Item",
                "Path":              "sites/pubsubpatterntesting/Lists/TranslationData_TST",
            })

        # UOM mapping rows
        for uom in t.get("uom_mappings", []):
            rows.append({
                "ReceiverName":      receiver,
                "MessageType":       msg_type,
                "TranslationSchema": "INTERNAL_TO_EXTERNAL",
                "TranslationName":   "uom",
                "ValueFrom":         uom["value_from"],
                "ValueTo":           uom["value_to"],
                "Created By":        created_by,
                "Modified By":       created_by,
                "Modified":          today,
                "Item Type":         "Item",
                "Path":              "sites/pubsubpatterntesting/Lists/TranslationData_TST",
            })

    return rows


# ────────────────────────────────────────────────────────────────────────────
# Output writers
# ────────────────────────────────────────────────────────────────────────────

TRANSLATION_COLUMNS = [
    "ReceiverName", "MessageType", "TranslationSchema", "TranslationName",
    "ValueFrom", "ValueTo", "Created By", "Modified By", "Modified",
    "Item Type", "Path",
]


def write_csv(rows: list[dict], path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TRANSLATION_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_excel(rows: list[dict], path: str):
    if not EXCEL_AVAILABLE:
        print("  [WARN] openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Translation Rows"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    row_fill = PatternFill("solid", fgColor="DDEEFF")
    alt_fill = PatternFill("solid", fgColor="EEF5FF")
    thin = Side(style="thin", color="BFBFBF")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    aln  = Alignment(horizontal="left", vertical="center", wrap_text=True)

    col_widths = [32, 18, 24, 30, 36, 16, 28, 28, 18, 10, 52]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    for ci, col in enumerate(TRANSLATION_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd

    for ri, row in enumerate(rows, 2):
        bg = row_fill if ri % 2 == 0 else alt_fill
        ws.row_dimensions[ri].height = 18
        for ci, col in enumerate(TRANSLATION_COLUMNS, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            c.fill = bg
            c.font = Font(name="Calibri", size=9)
            c.alignment = aln; c.border = brd

    ws.freeze_panes = "A2"
    wb.save(path)


# ────────────────────────────────────────────────────────────────────────────
# Summary printer
# ────────────────────────────────────────────────────────────────────────────

def print_summary(data: dict, yaml_block: str, translation_rows: list[dict],
                  out_yaml: str, out_xlsx: str, out_csv: str):
    ck = data["country_key"]
    nav = data["nav"]
    tx_enabled = [
        code for code, entry in nav["transaction_types"].items()
        if isinstance(entry, dict) and entry.get("enabled")
    ]
    print()
    print("=" * 70)
    print("  3PL ONBOARDING AUTOMATION — SUMMARY")
    print("=" * 70)
    print(f"  Country key        : {ck}")
    print(f"  NAV host           : {nav['host']}")
    print(f"  NAV company        : {nav['company']}")
    print(f"  Routing code       : {nav['routing_code']}")
    print(f"  Transaction types  : {', '.join(tx_enabled)}")
    print(f"  Receiver (BTP ID)  : {data['translation']['receiver_name']}")
    print(f"  Translation rows   : {len(translation_rows)}")
    print()
    print("  OUTPUT FILES")
    print(f"    Patched YAML     : {out_yaml}")
    print(f"    Translations XLS : {out_xlsx}")
    print(f"    Translations CSV : {out_csv}")
    print()
    print("  YAML BLOCK APPENDED")
    print("  " + "-" * 66)
    for line in yaml_block.strip().splitlines():
        print(f"  {line}")
    print("  " + "-" * 66)
    print()
    print("  TRANSLATION ROWS GENERATED")
    print("  " + "-" * 66)
    print(f"  {'ReceiverName':<32} {'TranslationName':<28} {'ValueFrom':<30} {'ValueTo'}")
    for r in translation_rows:
        print(f"  {r['ReceiverName']:<32} {r['TranslationName']:<28} {r['ValueFrom']:<30} {r['ValueTo']}")
    print("  " + "-" * 66)
    print()
    print("  NEXT STEPS")
    print("  1. Review the patched YAML diff before committing.")
    print("  2. Create a feature branch and commit the patched YAML.")
    print("  3. Import the CSV/Excel rows into the SharePoint Translation list.")
    print("  4. Raise a PR and notify the Mule reviewer.")
    print("=" * 70)
    print()


# ────────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Patch MuleSoft YAML + generate Translation rows from onboarding JSON."
    )
    parser.add_argument(
        "--input", "-i",
        default=None,
        help="Path to onboarding JSON (default: ../config/sample_onboarding_input.json)",
    )
    parser.add_argument(
        "--yaml", "-y",
        default=None,
        help="Path to MuleSoft YAML to patch (default: ../config/dev.yaml)",
    )
    parser.add_argument(
        "--out-yaml",
        default=None,
        help="Output path for patched YAML (default: <original>_patched.yaml)",
    )
    parser.add_argument(
        "--out-translations",
        default=None,
        help="Output path for translation Excel (default: translations_<country>.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be done without writing any files.",
    )
    args = parser.parse_args()

    # ── Resolve paths (script lives in scripts/, config in ../config/) ──────
    script_dir  = Path(__file__).parent
    config_dir  = script_dir.parent / "config"
    outputs_dir = script_dir.parent / "outputs"
    outputs_dir.mkdir(exist_ok=True)

    raw_input = args.input or str(config_dir / "sample_onboarding_input.json")
    raw_yaml  = args.yaml  or str(config_dir / "dev.yaml")

    input_path = Path(raw_input) if Path(raw_input).is_absolute() else Path(raw_input)
    yaml_path  = Path(raw_yaml)  if Path(raw_yaml).is_absolute()  else Path(raw_yaml)

    # ── Load input JSON ──────────────────────────────────────────────────────
    if not input_path.exists():
        print(f"[ERROR] Input JSON not found: {input_path}")
        sys.exit(1)

    with open(input_path, encoding="utf-8") as f:
        # Strip _comment keys before processing
        raw = json.load(f)

    data = {k: v for k, v in raw.items() if not k.startswith("_")}
    # Recursively strip _comment / _note keys from nested dicts
    def strip_comments(obj):
        if isinstance(obj, dict):
            return {k: strip_comments(v) for k, v in obj.items() if not k.startswith("_")}
        return obj
    data = strip_comments(data)

    # ── Validate ─────────────────────────────────────────────────────────────
    errors = validate(data)
    if errors:
        print("[ERROR] Validation failed — fix the following in your input JSON:")
        for e in errors:
            print(f"  * {e}")
        sys.exit(1)
    print("[OK] Input JSON validated.")

    # ── Check YAML exists ────────────────────────────────────────────────────
    if not yaml_path.exists():
        print(f"[ERROR] YAML file not found: {yaml_path}")
        sys.exit(1)

    # ── Build outputs ────────────────────────────────────────────────────────
    country_key = data["country_key"]

    out_yaml = (
        Path(args.out_yaml) if args.out_yaml
        else outputs_dir / f"{yaml_path.stem}_patched_{country_key}.yaml"
    )
    out_xlsx = (
        Path(args.out_translations) if args.out_translations
        else outputs_dir / f"translations_{country_key}.xlsx"
    )
    out_csv = out_xlsx.with_suffix(".csv")

    # ── Generate YAML block ──────────────────────────────────────────────────
    yaml_block = build_yaml_block(data)

    # ── Patch YAML content ───────────────────────────────────────────────────
    try:
        patched_content = patch_yaml(str(yaml_path), country_key, yaml_block)
    except ValueError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)

    # ── Generate translation rows ────────────────────────────────────────────
    translation_rows = build_translation_rows(data)

    # ── Print summary (always) ───────────────────────────────────────────────
    print_summary(data, yaml_block, translation_rows,
                  str(out_yaml), str(out_xlsx), str(out_csv))

    if args.dry_run:
        print("[DRY RUN] No files written.")
        return

    # ── Write outputs ────────────────────────────────────────────────────────
    with open(out_yaml, "w", encoding="utf-8") as f:
        f.write(patched_content)
    print(f"[OK] Patched YAML written  -> {out_yaml}")

    write_excel(translation_rows, str(out_xlsx))
    print(f"[OK] Translation Excel     -> {out_xlsx}")

    write_csv(translation_rows, str(out_csv))
    print(f"[OK] Translation CSV       -> {out_csv}")


if __name__ == "__main__":
    main()
