import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "3PL Inbound Onboarding"

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # dark navy  – sheet title / band headers
C_HEADER_FG   = "FFFFFF"
C_BTP_BG      = "2E75B6"   # medium blue – BTP section
C_SOL_BG      = "375623"   # dark green  – Solace section
C_MUL_BG      = "7B2C2C"   # dark red    – MuleSoft section
C_GEN_BG      = "4A4A4A"   # dark grey   – General section
C_COL_HDR_FG  = "FFFFFF"
C_Q_BG        = "D6E4F0"   # light blue  – BTP question rows
C_Q_SOL_BG   = "E2EFDA"   # light green
C_Q_MUL_BG   = "FCE4D6"   # light salmon
C_Q_GEN_BG   = "EDEDED"   # light grey
C_ALT_BTP    = "EBF3FB"
C_ALT_SOL    = "F0F7EB"
C_ALT_MUL    = "FEF2EC"
C_ALT_GEN    = "F5F5F5"
C_MANDATORY  = "FF0000"
C_ANSWER_BG  = "FFFDE7"   # pale yellow – answer cells

# ── Helper styles ───────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    t = Side(style="medium", color="4A4A4A")
    return Border(left=t, right=t, top=t, bottom=t)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = {
    1: 5,    # #
    2: 32,   # Question / Field
    3: 38,   # Answer / Value
    4: 18,   # Owner / Team
    5: 14,   # Status
    6: 30,   # Notes / Examples
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Freeze panes (keep header rows visible) ──────────────────────────────────
ws.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════
# Row 1 – Sheet title
# ═══════════════════════════════════════════════════════════════════════════
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "3PL → Mars Navision  |  Inbound Integration Onboarding  |  Requirements Capture"
c.fill = fill(C_HEADER_BG)
c.font = font(bold=True, color=C_HEADER_FG, size=13)
c.alignment = align(h="center")
ws.row_dimensions[1].height = 28

# ── Sub-title / instructions ─────────────────────────────────────────────
ws.merge_cells("A2:F2")
c = ws["A2"]
c.value = (
    "Complete all fields marked  ★ (mandatory).  "
    "Status values: [ ] Open  |  [~] In Progress  |  [✓] Confirmed.  "
    "Raise questions to the respective team owner before automation is triggered."
)
c.fill = fill("D9E1F2")
c.font = font(italic=True, size=9, color="2F2F2F")
c.alignment = align(h="left")
ws.row_dimensions[2].height = 18

# Row 3 – column headers
headers = ["#", "Question / Field", "Answer / Value", "Owner / Team", "Status", "Notes / Examples"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.fill = fill(C_HEADER_BG)
    c.font = font(bold=True, color=C_HEADER_FG, size=10)
    c.alignment = align(h="center")
    c.border = border()
ws.row_dimensions[3].height = 20

# ═══════════════════════════════════════════════════════════════════════════
# Data rows builder
# ═══════════════════════════════════════════════════════════════════════════
current_row = 4

def write_section_header(title, bg_color, rows_span=1):
    global current_row
    ws.merge_cells(f"A{current_row}:F{current_row}")
    c = ws[f"A{current_row}"]
    c.value = title
    c.fill = fill(bg_color)
    c.font = font(bold=True, color=C_COL_HDR_FG, size=11)
    c.alignment = align(h="left")
    c.border = thick_border()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

def write_row(num, question, answer="", owner="", status="[ ]", notes="", mandatory=False, *,
              q_fill_color=C_Q_BG, alt_fill=C_ALT_BTP, is_alt=False):
    global current_row
    bg = alt_fill if is_alt else q_fill_color
    mark = "  ★" if mandatory else ""
    data = [num, question + mark, answer, owner, status, notes]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=current_row, column=col, value=val)
        c.fill = fill(C_ANSWER_BG if col == 3 else bg)
        c.font = font(
            bold=(col == 1),
            color=(C_MANDATORY if mandatory and col == 2 else "1A1A1A"),
            size=9
        )
        c.alignment = align(h="center" if col in (1, 4, 5) else "left")
        c.border = border()
    ws.row_dimensions[current_row].height = 36
    current_row += 1

def spacer():
    global current_row
    ws.merge_cells(f"A{current_row}:F{current_row}")
    ws.row_dimensions[current_row].height = 8
    current_row += 1

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 0 – General Integration Details  (from schema)
# ═══════════════════════════════════════════════════════════════════════════
write_section_header("SECTION 0 — General Integration Details  (Interface Identification)", C_GEN_BG)

gen_rows = [
    (1,  "Interface ID",                         "", "Integration Lead", "[ ]", "E.g. INT-2024-042", True),
    (2,  "EA Reference Number",                  "", "Enterprise Arch",  "[ ]", "Format: EAXXXX",   True),
    (3,  "Source Application",                   "", "Project Team",     "[ ]", "E.g. 3PL WMS",     True),
    (4,  "Target Application",                   "", "Project Team",     "[ ]", "Mars Navision",     True),
    (5,  "Business Object",                      "", "Project Team",     "[ ]", "E.g. Stock Receipt",False),
    (6,  "Country / Market",                     "", "Project Team",     "[ ]", "ISO country code",  True),
    (7,  "Source Format",                        "", "3PL Team",         "[ ]", "xml / json / txt",  True),
    (8,  "Source Interface Type",                "", "3PL Team",         "[ ]", "api / idoc / file", True),
    (9,  "Target Format",                        "", "Navision Team",    "[ ]", "xml / json / txt",  True),
    (10, "Target Interface Type",                "", "Navision Team",    "[ ]", "api / idoc / file", True),
    (11, "Functional Description of Flow",       "", "Integration Lead", "[ ]", "Free text",         True),
    (12, "Volume — messages per time period",    "", "3PL Team",         "[ ]", "E.g. 500/day",      False),
    (13, "Size per message / event",             "", "3PL Team",         "[ ]", "E.g. 50 KB",        False),
    (14, "Peak Volume",                          "", "3PL Team",         "[ ]", "E.g. 2000 in 1 hr", False),
]
for i, r in enumerate(gen_rows):
    write_row(*r, q_fill_color=C_Q_GEN_BG, alt_fill=C_ALT_GEN, is_alt=(i % 2 == 1))

spacer()

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 – SAP BTP
# ═══════════════════════════════════════════════════════════════════════════
write_section_header("SECTION 1 — SAP BTP Team  (API Exposure & Application Registration)", C_BTP_BG)

btp_rows = [
    (1,  "Application name to register on BTP Developer Portal",
         "", "BTP Team", "[ ]", "Must match naming convention", True),
    (2,  "BTP Developer Portal URL (UAT)",
         "", "BTP Team", "[ ]", "https://…",                    True),
    (3,  "BTP Developer Portal URL (PROD)",
         "", "BTP Team", "[ ]", "https://…",                    True),
    (4,  "Subscription / Plan tier to select during registration",
         "", "BTP Team", "[ ]", "E.g. standard / premium",      True),
    (5,  "Client ID (User ID) issued after registration — ★ feeds Solace & Value Mapping",
         "", "BTP Team", "[ ]", "Auto-generated by BTP portal", True),
    (6,  "Client Secret / Password issued after registration",
         "", "BTP Team", "[ ]", "Store in Key Vault",           True),
    (7,  "Callback URL to configure on BTP Portal",
         "", "iHub Team","[ ]", "Provided by iHub",             True),
    (8,  "D3 Approver name / email (Human-in-the-loop approver)",
         "", "D3 / iHub", "[ ]","Person who approves the BTP request", True),
    (9,  "SLA for D3 approval (expected turnaround time)",
         "", "D3 / iHub", "[ ]","E.g. 2 business days",         False),
    (10, "SAP notification endpoint URL (UAT)",
         "", "BTP Team", "[ ]", "Endpoint SAP exposes for push notifications", True),
    (11, "SAP notification endpoint URL (PROD)",
         "", "BTP Team", "[ ]", "Endpoint SAP exposes for push notifications", True),
    (12, "Authentication method on SAP notification endpoint",
         "", "BTP Team", "[ ]", "OAuth2 / Basic / mTLS",        True),
    (13, "Event Grid topic name / connection string (UAT)",
         "", "BTP / Azure","[ ]","ASB event grid topic",         True),
    (14, "Event Grid topic name / connection string (PROD)",
         "", "BTP / Azure","[ ]","ASB event grid topic",         True),
    (15, "Azure Service Bus (ASB) queue / topic name (UAT)",
         "", "BTP / Azure","[ ]","Queue that receives BTP events",True),
    (16, "Azure Service Bus (ASB) queue / topic name (PROD)",
         "", "BTP / Azure","[ ]","Queue that receives BTP events",True),
    (17, "Value Mapping entry key/name to create for this partner",
         "", "iHub Team", "[ ]","Derived from Client ID",       True),
    (18, "iHub contact to whom User ID details should be communicated",
         "", "iHub Team", "[ ]","Name + email",                  True),
    (19, "Any IP allowlist / firewall rules needed on BTP side",
         "", "BTP Team", "[ ]","List of IPs to whitelist",      False),
    (20, "Certificate / mTLS details (if applicable)",
         "", "BTP Team", "[ ]","CN, expiry, issuer",            False),
]
for i, r in enumerate(btp_rows):
    write_row(*r, q_fill_color=C_Q_BG, alt_fill=C_ALT_BTP, is_alt=(i % 2 == 1))

spacer()

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 – Solace
# ═══════════════════════════════════════════════════════════════════════════
write_section_header("SECTION 2 — Solace Team  (Queue & Subscription Configuration)", C_SOL_BG)

sol_rows = [
    (1,  "Solace broker URL / host (UAT)",
         "", "Solace Team","[ ]","smf://… or https://…",        True),
    (2,  "Solace broker URL / host (PROD)",
         "", "Solace Team","[ ]","smf://… or https://…",        True),
    (3,  "VPN / Message VPN name (UAT)",
         "", "Solace Team","[ ]","E.g. mars-uat-vpn",           True),
    (4,  "VPN / Message VPN name (PROD)",
         "", "Solace Team","[ ]","E.g. mars-prod-vpn",          True),
    (5,  "Existing queue name to append the new subscription to",
         "", "Solace Team","[ ]","Must not touch existing subs", True),
    (6,  "New subscription rule / topic string to add",
         "", "Solace Team","[ ]","E.g. mars/3pl/{partnerID}/>", True),
    (7,  "3PL Partner ID  (= BTP Client ID from Section 1 Q5)",
         "", "Solace Team","[ ]","Auto-populated from BTP",      True),
    (8,  "1Nav Instance ID",
         "", "Solace Team","[ ]","E.g. NAV-EU-01",              True),
    (9,  "Country Code ISO",
         "", "Project Team","[ ]","E.g. FR / DE / NL",          True),
    (10, "Region ISO",
         "", "Project Team","[ ]","E.g. EU / APAC",             True),
    (11, "Confirm list of existing subscriptions on the queue (snapshot)",
         "", "Solace Team","[ ]","Attach export or paste here", True),
    (12, "Solace admin credentials / service account for automation",
         "", "Solace Team","[ ]","Store in Key Vault",          True),
    (13, "Solace SEMP API endpoint (for programmatic config)",
         "", "Solace Team","[ ]","https://broker/SEMP/v2/…",    False),
    (14, "Max message TTL / queue quota (if changed for this partner)",
         "", "Solace Team","[ ]","Default or custom value",     False),
    (15, "Dead-letter / DMQ queue name",
         "", "Solace Team","[ ]","If applicable",               False),
]
for i, r in enumerate(sol_rows):
    write_row(*r, q_fill_color=C_Q_SOL_BG, alt_fill=C_ALT_SOL, is_alt=(i % 2 == 1))

spacer()

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 – MuleSoft
# ═══════════════════════════════════════════════════════════════════════════
write_section_header("SECTION 3 — MuleSoft Team  (YAML Update, GitHub PR & Deployment)", C_MUL_BG)

mul_rows = [
    (1,  "GitHub repository URL for the MuleSoft project",
         "", "Mule Team","[ ]","https://github.com/org/repo",  True),
    (2,  "Base branch to branch from (dev / main)",
         "", "Mule Team","[ ]","E.g. dev",                     True),
    (3,  "Feature branch naming convention",
         "", "Mule Team","[ ]","E.g. feature/3pl-onboard-FR",  True),
    (4,  "Path to the YAML config file to be modified",
         "", "Mule Team","[ ]","src/main/resources/config.yml",True),
    (5,  "Country / Market being added to YAML (new entry)",
         "", "Project Team","[ ]","E.g. FR",                   True),
    (6,  "YAML key/section under which new country entry is added",
         "", "Mule Team","[ ]","E.g. countries.navision",      True),
    (7,  "GitHub agent / service account username (contributor identity)",
         "", "Mule Team","[ ]","Used to raise the PR",         True),
    (8,  "GitHub Personal Access Token / App credentials for automation",
         "", "Mule Team","[ ]","Store in Key Vault",           True),
    (9,  "MuleSoft developer / reviewer email to notify after PR raised",
         "", "Mule Team","[ ]","Name + email",                 True),
    (10, "Input JSON validation schema location / version",
         "", "Mule Team","[ ]","URL or file path",             True),
    (11, "MuleSoft API endpoint URL (UAT) — receives the onboarding payload",
         "", "Mule Team","[ ]","https://…",                    True),
    (12, "MuleSoft API endpoint URL (PROD)",
         "", "Mule Team","[ ]","https://…",                    True),
    (13, "Authentication method for MuleSoft API",
         "", "Mule Team","[ ]","OAuth2 / Basic / API Key",     True),
    (14, "Expected deployment environment after PR merge (DEV first)",
         "", "Mule Team","[ ]","DEV → UAT → PROD",             False),
    (15, "PR review SLA / approval turnaround",
         "", "Mule Team","[ ]","E.g. 1 business day",          False),
    (16, "Any mandatory PR checks / CI gates to pass",
         "", "Mule Team","[ ]","E.g. unit tests, lint",        False),
    (17, "Rollback procedure if DEV deployment fails",
         "", "Mule Team","[ ]","Revert PR / manual steps",     False),
]
for i, r in enumerate(mul_rows):
    write_row(*r, q_fill_color=C_Q_MUL_BG, alt_fill=C_ALT_MUL, is_alt=(i % 2 == 1))

spacer()

# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 – Connectivity Details  (from schema)
# ═══════════════════════════════════════════════════════════════════════════
write_section_header("SECTION 4 — Connectivity Details  (SFTP / API endpoints)", C_GEN_BG)

conn_rows = [
    # 1Nav Outbound
    (1,  "[1Nav O/B] SFTP Path — UAT",           "", "Navision Team","[ ]","Full SFTP path",       True),
    (2,  "[1Nav O/B] Filename Scheme — UAT",      "", "Navision Team","[ ]","E.g. RCPT_{date}.txt", True),
    (3,  "[1Nav O/B] SFTP Path — PROD",          "", "Navision Team","[ ]","Full SFTP path",       True),
    (4,  "[1Nav O/B] Filename Scheme — PROD",     "", "Navision Team","[ ]","E.g. RCPT_{date}.txt", True),
    # 1Nav Inbound
    (5,  "[1Nav I/B] Target URL — UAT",          "", "Navision Team","[ ]","https://nav-uat/api/…",True),
    (6,  "[1Nav I/B] User — UAT",                "", "Navision Team","[ ]","Service account",      True),
    (7,  "[1Nav I/B] Target URL — PROD",         "", "Navision Team","[ ]","https://nav-prd/api/…",True),
    (8,  "[1Nav I/B] User — PROD",               "", "Navision Team","[ ]","Service account",      True),
    # 3PL Outbound
    (9,  "[3PL O/B] Target URL — UAT",           "", "3PL Team",    "[ ]","https://3pl-uat/…",    True),
    (10, "[3PL O/B] User — UAT",                 "", "3PL Team",    "[ ]","Service account",      True),
    (11, "[3PL O/B] Target URL — PROD",          "", "3PL Team",    "[ ]","https://3pl-prd/…",    True),
    (12, "[3PL O/B] User — PROD",               "", "3PL Team",    "[ ]","Service account",      True),
]
for i, r in enumerate(conn_rows):
    write_row(*r, q_fill_color=C_Q_GEN_BG, alt_fill=C_ALT_GEN, is_alt=(i % 2 == 1))

spacer()

# ═══════════════════════════════════════════════════════════════════════════
# Legend row
# ═══════════════════════════════════════════════════════════════════════════
ws.merge_cells(f"A{current_row}:F{current_row}")
c = ws[f"A{current_row}"]
c.value = (
    "LEGEND:  ★ = Mandatory field     [ ] = Open     [~] = In Progress     [✓] = Confirmed     "
    "Column C (Answer/Value) highlighted in yellow — fill in with confirmed values."
)
c.fill = fill("FFF2CC")
c.font = font(italic=True, size=9, color="7F6000")
c.alignment = align(h="left")
c.border = thick_border()
ws.row_dimensions[current_row].height = 18

# ── Save ─────────────────────────────────────────────────────────────────
out = r"c:\Users\SoubhikMukhopadhyay\source\repos\ais-operations\OperationService2\GetKVSecretExpTime\3PL_Inbound_Onboarding_Requirements.xlsx"
wb.save(out)
print(f"Saved -> {out}")
