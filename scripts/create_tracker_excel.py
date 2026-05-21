"""
3PL Integration Automation - Requirements & Action Tracker
Generates: 3PL_Integration_Automation_Tracker.xlsx  (1 tab)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Integration Tracker"

# ── Palette ─────────────────────────────────────────────────────────────────
P = {
    "title_bg":    "1F4E79", "title_fg":    "FFFFFF",
    "col_bg":      "2E2E2E", "col_fg":      "FFFFFF",
    "btp_hdr":     "1565C0", "btp_row":     "DDEEFF", "btp_alt":     "EEF5FF",
    "sol_hdr":     "2E7D32", "sol_row":     "DDEFDD", "sol_alt":     "EEF8EE",
    "mul_hdr":     "7B1FA2", "mul_row":     "F3E5F5", "mul_alt":     "FAF0FF",
    "ans_bg":      "FFFDE7", "warn_bg":     "FFF3CD",
    "auto_fg":     "1B5E20", "manual_fg":   "B71C1C", "tbd_fg":      "E65100",
    "api_yes_fg":  "0D47A1", "api_no_fg":   "546E7A",
    "note_bg":     "F8F9FA",
}

def fill(c):   return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="1A1A1A", size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def aln(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def thick():
    t = Side(style="medium", color="333333")
    return Border(left=t, right=t, top=t, bottom=t)

# ── Column widths (A=1 … I=9) ────────────────────────────────────────────────
widths = {1:5, 2:32, 3:18, 4:14, 5:48, 6:36, 7:22, 8:44, 9:24}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "A4"

# ── Row 1 : Title ────────────────────────────────────────────────────────────
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "3PL Integration Automation  -  Requirements & Action Tracker  |  Inbound : 3PL -> Mars Navision"
c.fill = fill(P["title_bg"]); c.font = fnt(bold=True, color=P["title_fg"], size=13)
c.alignment = aln(h="center"); ws.row_dimensions[1].height = 28

# ── Row 2 : Sub-note ────────────────────────────────────────────────────────
ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = ("Status codes:  [AUTO] Automated  |  [MAN] Manual / Interim  |  [TBD] To Be Decided    "
           "API col:  YES / NO / N/A    "
           "Action col: who needs to act before this step can proceed")
c.fill = fill("D9E1F2"); c.font = fnt(italic=True, size=9, color="2F2F2F")
c.alignment = aln(); ws.row_dimensions[2].height = 16

# ── Row 3 : Column headers ────────────────────────────────────────────────────
col_headers = [
    "#", "Step / Activity", "Automation\nStatus", "API\nAvailable?",
    "API Endpoint(s)", "Payload / Input Required", "Owner / DRI",
    "Open Items / Blockers", "Action Required From"
]
for ci, h in enumerate(col_headers, 1):
    c = ws.cell(row=3, column=ci, value=h)
    c.fill = fill(P["col_bg"]); c.font = fnt(bold=True, color=P["col_fg"], size=9)
    c.alignment = aln(h="center"); c.border = brd()
ws.row_dimensions[3].height = 32

# ── Data writer ────────────────────────────────────────────────────────────────
ROW = 4

def section(title, bg_hdr):
    global ROW
    ws.merge_cells(f"A{ROW}:I{ROW}")
    c = ws[f"A{ROW}"]
    c.value = title; c.fill = fill(bg_hdr)
    c.font = fnt(bold=True, color="FFFFFF", size=11)
    c.alignment = aln(h="left"); c.border = thick()
    ws.row_dimensions[ROW].height = 22; ROW += 1

def row(num, step, auto_status, api_avail, endpoints, payload,
        owner, open_items, action_from, row_bg, alt_bg, is_alt=False):
    global ROW
    bg = alt_bg if is_alt else row_bg
    # auto_status colour
    if "AUTO" in auto_status.upper() or "AUTOM" in auto_status.upper():
        st_color = P["auto_fg"]
    elif "MAN" in auto_status.upper():
        st_color = P["manual_fg"]
    else:
        st_color = P["tbd_fg"]
    # api colour
    api_color = (P["api_yes_fg"] if "YES" in str(api_avail).upper()
                 else P["api_no_fg"] if "NO" in str(api_avail).upper()
                 else P["tbd_fg"])

    vals = [num, step, auto_status, api_avail, endpoints, payload, owner, open_items, action_from]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ROW, column=ci, value=v)
        if ci == 3:
            c.fill = fill(bg); c.font = fnt(bold=True, color=st_color, size=9)
        elif ci == 4:
            c.fill = fill(bg); c.font = fnt(bold=True, color=api_color, size=9)
        elif ci in (5, 6):
            c.fill = fill(P["ans_bg"]); c.font = fnt(size=9)
        elif ci == 8:
            c.fill = fill(P["warn_bg"]); c.font = fnt(italic=True, size=9, color="7F4000")
        elif ci == 9:
            c.fill = fill(bg); c.font = fnt(bold=True, size=9, color="4A235A")
        else:
            c.fill = fill(bg); c.font = fnt(size=9)
        c.alignment = aln(h="center" if ci in (1, 3, 4) else "left")
        c.border = brd()
    ws.row_dimensions[ROW].height = 70; ROW += 1

def spacer():
    global ROW
    ws.merge_cells(f"A{ROW}:I{ROW}")
    ws.row_dimensions[ROW].height = 7; ROW += 1

# ════════════════════════════════════════════════════════════════════════════
#  BTP
# ════════════════════════════════════════════════════════════════════════════
section("◆  BTP INTEGRATION STEPS", P["btp_hdr"])

btp = [
  (
    "BTP-1",
    "Application Creation for Product\n(Register new application in BTP Developer Portal)",
    "MANUAL\n(Interim)",
    "YES",
    "https://apim-management-host/odata/1.0/data.svc/APIMgmt.Applications",
    "TBD — BTP Team to share expected request payload\n"
    "(application name, description, product binding)",
    "BTP Team\n(D3 approval required)",
    "1. BTP Team to expose & document the API formally\n"
    "2. D3 Human-in-the-loop approval gate must be defined\n"
    "3. Decision needed: automate via API or keep manual process?\n"
    "4. Confirm whether D3 approval is per-onboarding or one-time",
    "BTP Team + D3 Approver"
  ),
  (
    "BTP-2",
    "Generate New ClientID & Secret\n(Subscription creation on BTP Portal)",
    "MANUAL\n(Interim)",
    "YES",
    "https://apim-management-host/odata/1.0/data.svc/APIMgmt.Applications\n"
    "(same base endpoint — sub-resource TBC)",
    "TBD — BTP Team to confirm exact endpoint and\nshare sample payload for subscription creation",
    "BTP Team",
    "1. BTP Team to confirm & share the exact API endpoint for subscription/credential generation\n"
    "2. Clarify: is ClientID auto-generated or parameterised?\n"
    "3. How are credentials securely returned to the caller?\n"
    "4. Does this step require a separate D3 approval?",
    "BTP Team"
  ),
  (
    "BTP-3",
    "Insert Values into ValueMapping Table\n"
    "(Upsert entry -> Save new version -> Deploy)",
    "AUTOMATED",
    "YES",
    "1. Upsert Entry:\n"
    "   https://effem-glb-ci-dev01-pr.integrationsuite.cfapps.us21.hana.ondemand.com"
    "/api/v1/UpsertValMaps\n\n"
    "2. Save as Version:\n"
    "   .../api/v1/ValueMappingDesigntimeArtifactSaveAsVersion\n\n"
    "3. Deploy:\n"
    "   .../api/v1/DeployValueMappingsDesigntimeArtifact",
    "TBD — Need Postman collection + sample payload from BTP Team.\n"
    "Expected fields: partnerID, countryCode, instanceID, mapping key/value pairs",
    "BTP Team\n(iHub to consume)",
    "1. BTP Team to share Postman collection with sample payloads for all 3 API calls\n"
    "2. Service Instance + Service Key needed — who provisions?\n"
    "3. BLOCKER: 3PL ValueMapping entries may not be editable once configured — CONFIRM\n"
    "4. Validate API with a test entry in DEV before wiring into automation\n"
    "5. Confirm OAuth scopes / credentials for calling these endpoints",
    "BTP Team"
  ),
]
for i, r in enumerate(btp):
    row(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8],
        P["btp_row"], P["btp_alt"], is_alt=(i % 2 == 1))

spacer()

# ════════════════════════════════════════════════════════════════════════════
#  SOLACE
# ════════════════════════════════════════════════════════════════════════════
section("◆  SOLACE INTEGRATION STEPS", P["sol_hdr"])

sol = [
  (
    "SOL-1",
    "Share BTP ClientID with Solace Team\n(Handoff: BTP-2 output -> Solace input)",
    "MANUAL",
    "N/A",
    "Output of BTP-2 (ClientID) is the 3PL Partner ID for Solace.\n"
    "No dedicated API — this is a data handoff step.",
    "N/A — data: ClientID string from BTP-2",
    "Azure / iHub Team",
    "1. Define formal handoff mechanism (email / ticket / shared vault?)\n"
    "2. Azure Team to document the handoff process between BTP and Solace teams\n"
    "3. Should ClientID be written to a shared Key Vault entry automatically?",
    "Azure Team"
  ),
  (
    "SOL-2",
    "Solace Prerequisites Capture\n"
    "1. Producer App Name\n"
    "2. Consumer App Name\n"
    "3. New Topic Subscriptions to add",
    "TBD",
    "TBD",
    "TBD — depends on Solace Event Portal API (see SOL-3)",
    "Producer App Name: ________________\n"
    "Consumer App Name: ________________\n"
    "New Topic Subscription strings: ________________",
    "Azure Team\n+ Solace Team",
    "1. Azure Team to define handoff process (who provides these names?)\n"
    "2. Confirm naming convention for Producer / Consumer App in Solace Event Portal\n"
    "3. Are these derived from the onboarding form or assigned by Solace admin?",
    "Azure Team + Solace Team"
  ),
  (
    "SOL-3",
    "Populate Event Portal - Design\n"
    "1. Create new Enum version (add country)\n"
    "2. Create new Event version (uses new Enum)\n"
    "3. Create new Producer App version + config\n"
    "4. Create new Consumer App version + subscriptions + config\n"
    "5. Submit Dev Promotion Request",
    "AUTOMATED",
    "YES",
    "Solace Cloud API Reference:\nhttps://api.solace.dev/cloud/reference\n\n"
    "API calls sequence:\n"
    "1. GET/POST Enumeration Version\n"
    "2. GET/POST Event Version\n"
    "3. GET/POST Application Version (Producer)\n"
    "4. GET/POST Application Version (Consumer)\n"
    "5. GET Model Event Brokers + POST App Promotion Request",
    "Per API call — to be defined.\n"
    "Core inputs needed:\n"
    "- countryCode (ISO)\n- regionISO\n- enumID / eventID / appIDs\n"
    "- subscription topic strings\n- broker environment ID",
    "Solace Team",
    "1. Solace Team to share API token / OAuth credentials for automation\n"
    "2. Provide IDs of existing Enum, Event, Producer App, Consumer App to GET before POST\n"
    "3. Confirm exact payload schema for each POST (share Postman collection)\n"
    "4. What is the Dev Promotion Request payload and approver?\n"
    "5. Is there a sandbox/dev Event Portal for testing automation before PROD?",
    "Solace Team"
  ),
  (
    "SOL-4",
    "Add Subscription Details to Existing Queue\n"
    "(Append new subscription rule; existing subs must remain intact)",
    "AUTOMATED",
    "YES",
    "TBD — Solace SEMP API (likely):\n"
    "PATCH /SEMP/v2/config/msgVpns/{vpn}/queues/{queue}/subscriptions\n"
    "(Confirm exact endpoint with Solace Team)",
    "Queue name: ________________\n"
    "New subscription topic string: ________________\n"
    "VPN name: ________________\n"
    "Broker host: ________________",
    "Solace Team",
    "1. Confirm SEMP API endpoint and auth method (basic / bearer)\n"
    "2. Share existing queue subscription snapshot (so automation validates before PATCH)\n"
    "3. Is there a 'dry-run' mode to validate without committing?\n"
    "4. Confirm: PATCH appends only — does not overwrite existing subscriptions?",
    "Solace Team"
  ),
  (
    "SOL-5",
    "Create Feature Branch (Git)\n"
    "(Branch from dev/main; path & file TBC)",
    "MANUAL\n(Interim — APIs available)",
    "YES",
    "GitHub / Azure DevOps API:\nPOST /repos/{org}/{repo}/git/refs\n"
    "(Git branch path and target file TBC with Solace & DevOps Team)",
    "Base branch: ________________\n"
    "New branch name: ________________\n"
    "Azure Service Principal: ________________",
    "Solace Team\n+ DevOps Team",
    "1. Azure Service Principal to be provisioned with 'create branch' rights on the repo\n"
    "2. Solace Team + DevOps to confirm repo URL and branch naming convention\n"
    "3. Which file(s) in the Git repo need to be updated for Solace config?\n"
    "4. Is this the same repo as MuleSoft or a separate Solace-config repo?",
    "Solace Team + DevOps Team"
  ),
  (
    "SOL-6",
    "Populate input.json from GitHub Branch\n"
    "(Ref: Confluence - Solace Event Broker GitHub Action Pipeline #RDP)",
    "TBD",
    "TBD",
    "Confluence reference:\nhttps://marsaoh.atlassian.net/wiki/spaces/IH/pages/"
    "4731961464/Solace+Event+Broker+GitHub+Action+Pipeline#RDP\n"
    "(Review actions currently executed manually)",
    "input.json fields TBC.\n"
    "Expected inputs from onboarding form:\n"
    "countryCode, regionISO, partnerID, subscriptionTopic, brokerEnv",
    "Solace Team\n+ DevOps Team",
    "1. Review Confluence page and list all manual steps — map each to an API call\n"
    "2. Define the full input.json schema (share sample file)\n"
    "3. What GitHub Action pipeline is triggered and what are its inputs?\n"
    "4. Who reviews / approves the populated input.json before pipeline runs?",
    "Solace Team + DevOps Team"
  ),
]
for i, r in enumerate(sol):
    row(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8],
        P["sol_row"], P["sol_alt"], is_alt=(i % 2 == 1))

spacer()

# ════════════════════════════════════════════════════════════════════════════
#  MULESOFT
# ════════════════════════════════════════════════════════════════════════════
section("◆  MULESOFT INTEGRATION STEPS", P["mul_hdr"])

mul = [
  (
    "MUL-1",
    "Create Feature Branch (Git)\n"
    "(Developer currently creates manually; APIs are available to automate)",
    "MANUAL\n(Interim — APIs available)",
    "YES",
    "GitHub API:\nPOST /repos/{org}/{repo}/git/refs\n\n"
    "Requires Azure Service Principal with 'create branch' rights.\n"
    "Confirm with Mule / DevOps Team.",
    "Base branch: ________________  (dev / main)\n"
    "New branch name: feature/3pl-{partnerID}-{countryCode}\n"
    "Repo URL: ________________\n"
    "Service Principal / PAT: ________________",
    "Sravani Kare\n(Contractor)\n+ Mule Team",
    "1. Azure Service Principal (or GitHub App) to be provisioned for automation\n"
    "2. Confirm repo URL and branch naming convention with Mule Team\n"
    "3. Is base branch 'dev' or 'main'? Confirm environment strategy\n"
    "4. Who merges the feature branch after validation?",
    "Mule Team\n(Sravani Kare)"
  ),
  (
    "MUL-2",
    "Update YAML – Add Country-Specific Configuration\n"
    "(Automation reads onboarding JSON -> patches existing YAML -> commits to feature branch)",
    "AUTOMATED",
    "TBD",
    "TBD — Sravani Kare to share API or confirm:\n"
    "Option A: Direct Git file update via GitHub API\n"
    "   PUT /repos/{org}/{repo}/contents/{path}\n"
    "Option B: MuleSoft management API endpoint\n"
    "   (Sravani to share endpoint details)",
    "Input: onboarding JSON (see schema below)\n"
    "Output: updated YAML with new country block appended.\n\n"
    "Country-specific YAML fields to define:\n"
    "- countryCode\n- navisionEndpointUAT\n- navisionEndpointPROD\n"
    "- partnerID\n- solaceSubscriptionTopic\n- [others TBC with Mule Team]\n\n"
    "NOTE: Sample YAML/JSON to be provided by Mule Team;\n"
    "placeholder sample will be used in interim.",
    "Sravani Kare\n(Contractor)\n+ Mule Team",
    "1. Sravani Kare to share: (a) API endpoint or (b) confirm Git API approach\n"
    "2. Share the existing YAML file structure + sample with one country entry\n"
    "3. Define the exact YAML key path where new country block is inserted\n"
    "4. Confirm: which fields in the onboarding JSON map to which YAML keys?\n"
    "5. After YAML update, who validates DEV deployment before UAT promotion?\n"
    "6. Notification email: send PR link + diff to Mule reviewer after commit",
    "Mule Team\n(Sravani Kare)"
  ),
  (
    "MUL-3",
    "Validate Input JSON\n"
    "(Schema validation before any API calls are triggered)",
    "AUTOMATED",
    "N/A\n(Internal)",
    "N/A — internal validation step.\n"
    "Validation logic runs against JSON schema before calling BTP / Solace / Mule APIs.",
    "JSON schema to validate:\n"
    "- Interfaces Description fields\n- Connectivity Details\n"
    "- Solace fields (partnerID, countryCode, regionISO, subscriptionRules)\n"
    "- Non-functional requirements\n\n"
    "Schema source: onboarding Excel / intake form",
    "iHub / Integration Team",
    "1. Finalise mandatory vs optional field list across all sections\n"
    "2. Define validation rules (regex for ISO codes, URL format checks, etc.)\n"
    "3. What happens on validation failure — halt and notify requester?",
    "Integration Team"
  ),
  (
    "MUL-4",
    "Raise Pull Request & Notify Mule Reviewer\n"
    "(After YAML update committed to feature branch)",
    "AUTOMATED",
    "YES",
    "GitHub API:\nPOST /repos/{org}/{repo}/pulls\n\n"
    "Email notification via:\nSendGrid / Azure Communication Services / SMTP relay\n"
    "(confirm channel with team)",
    "PR payload:\n"
    "- title: 'feat: 3PL onboarding {partnerID} - {countryCode}'\n"
    "- head: feature branch name\n- base: dev\n"
    "- body: summary of changes + diff link\n\n"
    "Email: PR URL + YAML diff + deployment checklist",
    "Automation\n-> Sravani Kare\n(Reviewer)",
    "1. Confirm GitHub reviewer / team to assign on PR\n"
    "2. Confirm notification channel: email / Teams / Slack?\n"
    "3. What is the expected review + merge SLA?\n"
    "4. Is DEV auto-deployed on PR merge or does it need manual trigger?",
    "Mule Team\n(Sravani Kare)"
  ),
]
for i, r in enumerate(mul):
    row(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8],
        P["mul_row"], P["mul_alt"], is_alt=(i % 2 == 1))

spacer()

# ════════════════════════════════════════════════════════════════════════════
#  SUMMARY: Open Actions by Team
# ════════════════════════════════════════════════════════════════════════════
ws.merge_cells(f"A{ROW}:I{ROW}")
c = ws[f"A{ROW}"]
c.value = "OPEN ACTIONS SUMMARY  -  Grouped by Team  (auto-derived from tracker above)"
c.fill = fill("263238"); c.font = fnt(bold=True, color="FFFFFF", size=11)
c.alignment = aln(h="left"); c.border = thick()
ws.row_dimensions[ROW].height = 22; ROW += 1

summary = [
    ("BTP Team",
     "Share Postman collection + payloads for all 3 ValueMapping API calls.\n"
     "Confirm exact endpoint for ClientID/Secret generation.\n"
     "Clarify D3 approval process and decision: automate App Creation or keep manual?\n"
     "Provision Service Instance + Service Key for accessing BTP APIs.\n"
     "Confirm whether 3PL ValueMapping entries are editable once configured."),
    ("Solace Team",
     "Share API credentials / OAuth token for Event Portal API automation.\n"
     "Provide IDs of existing Enum, Event, Producer App, Consumer App objects.\n"
     "Confirm SEMP API endpoint and auth method for queue subscription PATCH.\n"
     "Share existing queue subscription snapshot before automation runs.\n"
     "Provide input.json schema and sample file (ref: Confluence RDP page).\n"
     "Confirm Solace sandbox/dev Event Portal is available for testing."),
    ("Mule Team\n(Sravani Kare)",
     "Share: API endpoint OR confirm GitHub API approach for YAML update (MUL-2).\n"
     "Provide existing YAML file with at least one country entry as sample.\n"
     "Define exact YAML key path for new country block insertion.\n"
     "Confirm JSON -> YAML field mapping.\n"
     "Confirm reviewer GitHub handle to assign on auto-raised PR.\n"
     "Confirm notification channel (email / Teams) and reviewer email address."),
    ("Azure / DevOps Team",
     "Provision Azure Service Principal with 'create branch' rights on Solace + Mule repos.\n"
     "Define and document BTP -> Solace ClientID handoff mechanism.\n"
     "Confirm Key Vault path for storing ClientID, secrets, and service account credentials.\n"
     "Confirm notification relay (SendGrid / Azure Communication Services / SMTP)."),
    ("Integration / iHub Team",
     "Define complete JSON schema for onboarding intake form (mandatory + optional fields).\n"
     "Define validation rules (ISO code regex, URL format, field constraints).\n"
     "Create Value Mapping entry using ClientID from BTP-2 — coordinate with BTP Team.\n"
     "Communicate User/ClientID to Solace Team after BTP provisioning.\n"
     "Define end-to-end orchestration sequence and error-handling / rollback strategy."),
]

team_colors = ["BBDEFB", "C8E6C9", "E1BEE7", "FFE0B2", "E0E0E0"]
team_dark   = ["1565C0", "2E7D32", "7B1FA2", "E65100", "424242"]

for i, (team, actions) in enumerate(summary):
    ws.merge_cells(f"A{ROW}:B{ROW}")
    tc = ws[f"A{ROW}"]
    tc.value = team; tc.fill = fill(team_colors[i])
    tc.font = fnt(bold=True, color=team_dark[i], size=10); tc.alignment = aln()
    tc.border = brd()

    ws.merge_cells(f"C{ROW}:I{ROW}")
    ac = ws[f"C{ROW}"]
    ac.value = actions; ac.fill = fill(P["warn_bg"])
    ac.font = fnt(size=9, italic=True); ac.alignment = aln()
    ac.border = brd()
    ws.row_dimensions[ROW].height = 80; ROW += 1

# ── Legend ────────────────────────────────────────────────────────────────────
spacer()
ws.merge_cells(f"A{ROW}:I{ROW}")
lc = ws[f"A{ROW}"]
lc.value = ("LEGEND:   Automation Status -> AUTOMATED (green) | MANUAL/Interim (red) | TBD (orange)    "
            "API col -> YES (blue) | NO / N/A (grey)    "
            "Yellow cells = values to be filled in    "
            "Orange cells = open items / blockers requiring action")
lc.fill = fill("F8F9FA"); lc.font = fnt(italic=True, size=8, color="546E7A")
lc.alignment = aln(); lc.border = brd()
ws.row_dimensions[ROW].height = 16

# ── Save ──────────────────────────────────────────────────────────────────────
out = (r"c:\Users\SoubhikMukhopadhyay\source\repos\ais-operations"
       r"\OperationService2\GetKVSecretExpTime"
       r"\3PL_Integration_Automation_Tracker.xlsx")
wb.save(out)
print(f"Saved -> {out}")
